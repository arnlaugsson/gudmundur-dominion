#!/usr/bin/env python3
"""
Download a public Google Spreadsheet (or read a local XLSX file) and parse
the "Oll spil" sheet into a unified games array for dominion_data.json.

Usage:
    python3 scripts/sync_spreadsheet.py                    # download from Google
    python3 scripts/sync_spreadsheet.py path/to/local.xlsx # use local file
"""

import datetime
import json
import sys
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl

SPREADSHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "14f5TW05_jpnWGf6d3cAO7ChSrHPZl0dIbxyIO3qqe2A/export?format=xlsx"
)
ROOT = Path(__file__).resolve().parent.parent
JSON_FILE = ROOT / "data" / "dominion_data.json"
DOWNLOAD_PATH = ROOT / "data" / "spreadsheet.xlsx"

SHEET_NAME = "Öll spil"

# The sheet's row layout is NOT hardcoded. Rows are located at runtime from the
# labels in column A (see detect_layout) so that inserting or removing rows in
# the Google Sheet — e.g. adding the Prophecy row, or a second Project slot —
# never shifts a field into the wrong bucket. ROW_GAME_NUM is the one fixed
# anchor: game numbers always live on the second row and drive column discovery.
ROW_GAME_NUM = 2

# Each key maps to one label in column A (matched case-insensitively). Labels
# with a trailing hint like "Kingdom spil (G)" are matched by prefix.
SECTION_LABELS = {
    "date": ("Dagsetning", "exact"),
    "players": ("Spilarar", "exact"),
    "kingdom": ("Kingdom", "prefix"),
    # Row labelled "Auka spil" is an 11th kingdom-card slot, not a section of
    # its own. It is folded into the kingdom block in detect_layout().
    "kingdom_extra": ("Auka spil", "exact"),
    "events": ("Event", "prefix"),
    "landmarks": ("Landmark", "prefix"),
    "projects": ("Project", "exact"),
    "way": ("Way", "exact"),
    "allies": ("Allies", "exact"),
    "trait": ("Trait", "exact"),
    "prophecy": ("Prophecy", "exact"),
    "expansions": ("Sett", "exact"),
    "victory": ("Sigurtegund", "exact"),
    "scores": ("Sæti", "exact"),
    "avg": ("Meðalstig", "exact"),
}
# Sections that must exist for a parse to be trustworthy. "prophecy" is optional
# so older exports made before the Prophecy row existed still parse.
REQUIRED_SECTIONS = ("players", "kingdom", "expansions", "victory", "scores")
# How far down column A to scan for labels.
LABEL_SCAN_ROWS = 80


def detect_layout(ws):
    """Locate each section by its column-A label and return a layout dict.

    Multi-row sections (players, kingdom, expansions, scores, ...) are returned
    as an inclusive (start, end) row range that runs from the section's label
    up to the row before the next label — so the block auto-resizes when rows
    are inserted. Single-value sections (trait, victory) return an int row.
    Raises RuntimeError if a required label is missing (fail fast, don't guess).
    """
    # All labelled rows in column A, in order: (row, lowercased text).
    labelled = []
    for row in range(1, LABEL_SCAN_ROWS + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and val.strip():
            labelled.append((row, val.strip().lower()))
    label_rows = [row for row, _ in labelled]

    def find(label, mode):
        target = label.lower()
        for row, text in labelled:
            if (text == target) if mode == "exact" else text.startswith(target):
                return row
        return None

    found = {key: find(label, mode) for key, (label, mode) in SECTION_LABELS.items()}

    missing = [k for k in REQUIRED_SECTIONS if found.get(k) is None]
    if missing:
        raise RuntimeError(
            f"Could not locate required sheet sections {missing} by their labels "
            f"in column A. Detected labels: {[t for _, t in labelled]}"
        )

    def next_label_row(row):
        later = [r for r in label_rows if r > row]
        return min(later) if later else row

    def block(row):
        return (row, next_label_row(row) - 1)

    # "Auka spil" holds an 11th kingdom card. Because it carries its own label
    # it would otherwise close the kingdom block early and the card would be
    # dropped, so extend the block over it when it sits directly below.
    kingdom_block = block(found["kingdom"])
    extra = found.get("kingdom_extra")
    if extra is not None and extra == kingdom_block[1] + 1:
        kingdom_block = (kingdom_block[0], block(extra)[1])

    layout = {
        "game_num": ROW_GAME_NUM,
        "date": found["date"] or 1,
        "players": block(found["players"]),
        "kingdom": kingdom_block,
        "events": block(found["events"]) if found["events"] else None,
        "landmarks": block(found["landmarks"]) if found["landmarks"] else None,
        "projects": block(found["projects"]) if found["projects"] else None,
        "way": block(found["way"]) if found["way"] else None,
        "allies": block(found["allies"]) if found["allies"] else None,
        "trait": found["trait"],
        "prophecy": block(found["prophecy"]) if found["prophecy"] else None,
        "expansions": block(found["expansions"]),
        "victory": found["victory"],
        "scores": block(found["scores"]),
        "avg": found["avg"],
    }
    return layout


def download_spreadsheet(dest):
    """Download the public spreadsheet as XLSX."""
    print(f"Downloading spreadsheet to {dest} ...")
    try:
        urllib.request.urlretrieve(SPREADSHEET_URL, str(dest))
    except urllib.error.URLError as e:
        print(f"Error: Failed to download spreadsheet: {e}")
        sys.exit(1)
    print("Download complete.")
    return dest


def normalize_date(raw):
    """Convert date string like '2025.10.17' or '2025.10.17.' to 'YYYY-MM-DD'."""
    if raw is None:
        return None
    if isinstance(raw, datetime.datetime):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip().rstrip(".")
    if not s:
        return None
    parts = s.split(".")
    if len(parts) == 3:
        return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    return s


def is_valid_place(val):
    """Check if a cell value is a valid place number (not a datetime)."""
    if isinstance(val, datetime.datetime):
        return False
    if isinstance(val, (int, float)):
        return True
    return False


def read_cell(ws, row, col):
    """Read a cell value, returning None for empty-ish values."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, str):
        stripped = val.strip()
        if stripped == "" or stripped.lower() == "x":
            return None
        return stripped
    return val


def read_string_cell(ws, row, col):
    """Read a cell as a string, returning None for empty/placeholder values."""
    val = read_cell(ws, row, col)
    if val is None:
        return None
    if isinstance(val, str):
        return val
    return None


def find_game_columns(ws):
    """Scan row 2 for game number columns. Returns list of (game_num, col)."""
    game_cols = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=ROW_GAME_NUM, column=col).value
        if val is None:
            continue
        if not isinstance(val, (int, float)):
            continue
        num = int(val)
        if num < 1 or num > 9999:
            continue
        # Skip column 1 and 2 which are label columns
        if col <= 2:
            continue
        game_cols.append((num, col))
    return game_cols


def find_place_column(ws, col, players_range):
    """Find which column has the place numbers (normally col+2, but may be shifted
    if extra player names were entered in cols between name and place)."""
    for offset in (2, 3, 4):
        # Check if any player row has a valid place number in this column
        for row in range(players_range[0], players_range[1] + 1):
            val = ws.cell(row=row, column=col + offset).value
            if is_valid_place(val):
                return col + offset
    return col + 2  # default


def parse_players_and_places(ws, col, players_range):
    """Parse player names and places from the players block."""
    place_col = find_place_column(ws, col, players_range)
    players = []
    for row in range(players_range[0], players_range[1] + 1):
        name = read_string_cell(ws, row, col)
        if name is None:
            continue
        name = normalize_player_name(name)
        place_val = ws.cell(row=row, column=place_col).value
        place = int(place_val) if is_valid_place(place_val) else None
        players.append({"name": name, "place": place})
    return players


def parse_kingdom(ws, col, kingdom_range):
    """Parse kingdom cards from the kingdom block (card in col, expansion in col+1)."""
    cards = []
    for row in range(kingdom_range[0], kingdom_range[1] + 1):
        card_name = read_string_cell(ws, row, col)
        if card_name is None:
            continue
        expansion = read_string_cell(ws, row, col + 1)
        cards.append({"card": card_name, "expansion": expansion})
    return cards


def parse_string_rows(ws, col, rows):
    """Parse string values from a list of rows, filtering out None."""
    values = []
    for row in rows:
        val = read_string_cell(ws, row, col)
        if val is not None:
            values.append(val)
    return values


# Source-expansion tags (e.g. 'Renaissance') sit in col+1 of single-name rows
# to mark which expansion a project/way/trait comes from. They look like names
# but should not be treated as project names.
KNOWN_EXPANSION_TAGS = frozenset({
    'dominion', 'intrigue', 'seaside', 'alchemy', 'prosperity',
    'prospoerity', 'cornucopia', 'guilds', 'cornucopia & guilds',
    'hinterlands', 'dark ages', 'adventures', 'empires', 'nocturne',
    'renaissance', 'menagerie', 'allies', 'plunder', 'rising sun',
})


def is_expansion_tag(val):
    return isinstance(val, str) and val.strip().lower() in KNOWN_EXPANSION_TAGS


def parse_projects(ws, col, projects_range):
    """Parse the project block.

    A game may have more than one project, written either into extra rows of
    the block or into the next column of the same row. col+1 historically holds
    the source-expansion tag (e.g. 'Renaissance') when there is only one
    project, so we read every cell and drop expansion tags — every project
    survives regardless of which cell the user happened to use.
    """
    names = []
    for row in range(projects_range[0], projects_range[1] + 1):
        for offset in (0, 1, 2):
            val = read_string_cell(ws, row, col + offset)
            if val is None or is_expansion_tag(val):
                continue
            names.append(val)
    return names


def parse_traits(ws, col, trait_row):
    """Parse the trait row.

    Each trait may be paired with the kingdom-card pile it's attached to:
      - col+0: trait name (e.g., 'Rich')
      - col+1: source-expansion tag (e.g., 'Plunder') — informational, ignored here
      - col+2: attached kingdom card (e.g., 'Quartermaster') — may be empty for
        older games where the association wasn't recorded.
    Returns a list with at most one entry of shape {'name': str, 'card'?: str}.
    """
    name = read_string_cell(ws, trait_row, col)
    if name is None:
        return []
    card = read_string_cell(ws, trait_row, col + 2)
    if card:
        card = card.strip()
        if card.startswith("(") and card.endswith(")"):
            card = card[1:-1].strip()
        if not card:
            card = None
    entry = {"name": name}
    if card:
        entry["card"] = card
    return [entry]


def find_score_column(ws, col, scores_range):
    """Find which column has score numbers (normally col+2, may be shifted)."""
    for offset in (2, 3, 4):
        for row in range(scores_range[0], scores_range[1] + 1):
            val = ws.cell(row=row, column=col + offset).value
            if isinstance(val, (int, float)):
                return col + offset
    return col + 2


def parse_scores(ws, col, scores_range):
    """Parse the scores block. Returns list of {name, score}."""
    score_col = find_score_column(ws, col, scores_range)
    scores = []
    for row in range(scores_range[0], scores_range[1] + 1):
        name = read_string_cell(ws, row, col)
        if name is None:
            continue
        name = normalize_player_name(name)
        score_val = ws.cell(row=row, column=score_col).value
        score = float(score_val) if isinstance(score_val, (int, float)) else None
        scores.append({"name": name, "score": score})
    return scores


def parse_avg_score(ws, col, avg_row, scores_range):
    """Parse average score from the avg row."""
    if avg_row is None:
        return None
    score_col = find_score_column(ws, col, scores_range)
    val = ws.cell(row=avg_row, column=score_col).value
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    return None


PLAYER_NAME_FIXES = {
    "Hallgríur": "Hallgrímur",
}

EXPANSION_FIXES = {
    "Cornucopia": "Cornucopia & Guilds",
    "Prospoerity": "Prosperity",
    "Rising sun": "Rising Sun",
    "Dark ages": "Dark Ages",
}


def normalize_player_name(raw):
    """Fix known player name typos."""
    if raw is None:
        return None
    return PLAYER_NAME_FIXES.get(raw, raw)


def normalize_expansion(raw):
    """Fix known expansion name typos and inconsistencies."""
    if raw is None:
        return None
    return EXPANSION_FIXES.get(raw, raw)


def normalize_victory_type(raw):
    """Normalize victory type string."""
    if raw is None:
        return None
    mapping = {
        "province": "Province",
        "provinces": "Province",
        "provincce": "Province",
        "colony": "Colony",
        "colonies": "Colony",
        "supply piles": "Supply piles",
        "supply": "Supply piles",
    }
    return mapping.get(raw.lower(), raw)


def build_results(player_places, score_list):
    """
    Merge places from rows 3-7 with scores from rows 35-38 by matching
    player names. Sort by place. Detect ties.
    """
    score_map = {s["name"]: s["score"] for s in score_list}

    results = []
    for pp in player_places:
        if pp["place"] is None:
            continue
        results.append({
            "place": pp["place"],
            "name": pp["name"],
            "tied_with": None,
            "score": score_map.get(pp["name"]),
        })

    results.sort(key=lambda r: r["place"])

    # Detect ties: players with same place number
    place_groups = {}
    for r in results:
        p = r["place"]
        if p not in place_groups:
            place_groups[p] = []
        place_groups[p].append(r["name"])

    for r in results:
        group = place_groups[r["place"]]
        if len(group) > 1:
            others = [name for name in group if name != r["name"]]
            r["tied_with"] = others

    return results


def block_rows(rng):
    """Turn an inclusive (start, end) block into a row range, or [] if absent."""
    if not rng:
        return []
    return range(rng[0], rng[1] + 1)


def parse_game(ws, game_num, col, layout):
    """Parse a single game from its 3-column group using the detected layout."""
    date_raw = ws.cell(row=layout["date"], column=col).value
    date = normalize_date(date_raw)
    location = read_string_cell(ws, layout["game_num"], col + 1)

    player_places = parse_players_and_places(ws, col, layout["players"])

    # Skip incomplete games: no players or no valid places
    if not player_places:
        return None

    has_places = any(pp["place"] is not None for pp in player_places)

    kingdom = parse_kingdom(ws, col, layout["kingdom"])
    events = parse_string_rows(ws, col, block_rows(layout["events"]))
    landmarks = parse_string_rows(ws, col, block_rows(layout["landmarks"]))
    projects = parse_projects(ws, col, layout["projects"]) if layout["projects"] else []
    ways = parse_string_rows(ws, col, block_rows(layout["way"]))
    allies = parse_string_rows(ws, col, block_rows(layout["allies"]))
    traits = parse_traits(ws, col, layout["trait"]) if layout["trait"] else []
    prophecy = parse_string_rows(ws, col, block_rows(layout["prophecy"]))
    expansions = [normalize_expansion(e) for e in parse_string_rows(ws, col, block_rows(layout["expansions"]))]
    victory_type = normalize_victory_type(read_string_cell(ws, layout["victory"], col))

    score_list = parse_scores(ws, col, layout["scores"])

    # If no places but we have scores, infer places from scores (highest = 1st)
    if not has_places and score_list:
        scored = [s for s in score_list if s["score"] is not None]
        if scored:
            scored.sort(key=lambda s: s["score"], reverse=True)
            score_to_place = {}
            place = 1
            for i, s in enumerate(scored):
                if i > 0 and s["score"] == scored[i - 1]["score"]:
                    score_to_place[s["name"]] = score_to_place[scored[i - 1]["name"]]
                else:
                    score_to_place[s["name"]] = place
                place += 1
            for pp in player_places:
                if pp["name"] in score_to_place:
                    pp["place"] = score_to_place[pp["name"]]
            has_places = any(pp["place"] is not None for pp in player_places)

    # Games with no places and no scores are still included (with empty results)
    # so they count toward player stats like "biggest game"
    avg_score = parse_avg_score(ws, col, layout["avg"], layout["scores"])

    results = build_results(player_places, score_list)
    players = [pp["name"] for pp in player_places]

    return {
        "game_num": game_num,
        "date": date,
        "location": location,
        "players": players,
        "results": results,
        "kingdom": kingdom,
        "events": events,
        "landmarks": landmarks,
        "projects": projects,
        "ways": ways,
        "allies": allies,
        "traits": traits,
        "prophecy": prophecy,
        "expansions": expansions,
        "victory_type": victory_type,
        "avg_score": avg_score,
    }


def parse_spreadsheet(xlsx_path):
    """Parse all games from the spreadsheet."""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(
            f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[SHEET_NAME]
    layout = detect_layout(ws)
    print(f"Detected layout: {layout}")
    game_columns = find_game_columns(ws)
    print(f"Found {len(game_columns)} game columns in spreadsheet")

    games = []
    skipped = 0

    for game_num, col in game_columns:
        game = parse_game(ws, game_num, col, layout)
        if game is None:
            skipped += 1
            continue
        games.append(game)

    games.sort(key=lambda g: g["game_num"])

    print(f"Parsed {len(games)} games, skipped {skipped} incomplete")
    return games, skipped



# --- Card-list repair -------------------------------------------------------
# Only the "Öll spil" sheet is parsed here; the per-expansion card sheets
# (Base, Intrigue, ... Promo) are not. That means `cards` is carried over from
# the original one-off import on every sync, and corrections made in those
# sheets never reach the site on their own. Until the card sheets are parsed
# too, known problems are repaired below so a sync converges on correct data.

# Misspellings in the card list, mapped to the spelling the card sheets use.
CARD_NAME_FIXES = {
    "Transmongrify": "Transmogrify",
}

# Entries that duplicate a real card under the wrong expansion. Keyed by
# (name, expansion) so the correct entry with the same name survives.
CARD_DROPS = {
    ("Transmogrify", "Intrigue"),
}


def card_names_used(game):
    """Every card/event/landmark/... name a game references."""
    names = []
    for entry in game.get("kingdom") or []:
        if isinstance(entry, dict) and entry.get("card"):
            names.append(entry["card"])
    for key in ("events", "landmarks", "projects", "ways", "allies", "prophecy"):
        for value in game.get(key) or []:
            if isinstance(value, str):
                names.append(value)
    for trait in game.get("traits") or []:
        if not isinstance(trait, dict):
            continue
        if trait.get("name"):
            names.append(trait["name"])
        # The pile a trait is attached to is a kingdom pile in that game.
        if trait.get("card"):
            names.append(trait["card"])
    return names


def count_card_usage(games):
    """Map lowercased card name -> number of games it appears in.

    Matched case-insensitively: "Öll spil" and the card list disagree on
    capitalisation for many names (e.g. "Wishing well" vs "Wishing Well"),
    and a case-sensitive count silently reports 0 for those.
    """
    counts = {}
    for game in games:
        for name in {n.lower() for n in card_names_used(game)}:
            counts[name] = counts.get(name, 0) + 1
    return counts


def repair_cards(cards, games):
    """Drop bogus entries, fix known misspellings, merge duplicates, recount use."""
    counts = count_card_usage(games)
    repaired = []
    by_name = {}

    for card in cards:
        if (card.get("name"), card.get("expansion")) in CARD_DROPS:
            print(f"Dropped bogus card entry: {card['name']} ({card['expansion']})")
            continue

        card = dict(card)
        fixed = CARD_NAME_FIXES.get(card.get("name"))
        if fixed:
            print(f"Renamed card: {card['name']} -> {fixed}")
            card["name"] = fixed

        key = card["name"].lower()
        existing = by_name.get(key)
        if existing is not None:
            # Same card listed twice; fill blanks from the later entry rather
            # than discarding data (e.g. Shelters, listed with and without cost).
            for field, value in card.items():
                if existing.get(field) in (None, "") and value not in (None, ""):
                    existing[field] = value
            print(f"Merged duplicate card entry: {card['name']}")
            continue

        by_name[key] = card
        repaired.append(card)

    for card in repaired:
        card["times_used"] = counts.get(card["name"].lower(), 0)

    return repaired


def update_json(games, upcoming_count):
    """Update dominion_data.json with new games array."""
    if JSON_FILE.exists():
        with open(JSON_FILE, encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = {}

    # Construct a new dict instead of mutating the loaded one
    new_data = {k: v for k, v in data.items() if k != "legacy_games"}
    new_data["games"] = games
    if new_data.get("cards"):
        new_data["cards"] = repair_cards(new_data["cards"], games)
    new_data["last_updated"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    new_data["upcoming_games"] = upcoming_count

    if "legacy_games" in data:
        print("Removed 'legacy_games' key")

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(new_data, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"Updated {JSON_FILE} with {len(games)} games")


def main():
    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1])
        if not xlsx_path.exists():
            print(f"Error: File not found: {xlsx_path}")
            sys.exit(1)
    else:
        xlsx_path = DOWNLOAD_PATH
        download_spreadsheet(xlsx_path)

    games, skipped = parse_spreadsheet(xlsx_path)
    # Only count games beyond the current max completed game as pending
    max_completed = max((g["game_num"] for g in games if g["results"]), default=0)
    pending = sum(1 for g in games if g["game_num"] > max_completed and g["players"] and not g["results"])
    update_json(games, pending)


if __name__ == "__main__":
    main()
