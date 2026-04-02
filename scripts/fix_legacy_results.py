#!/usr/bin/env python3
"""
One-time script to fix legacy game results (games #1-49) in dominion_data.json.

The original JSON parsing scrambled name↔place pairings. This script reads
the correct data from the Excel spreadsheet and patches the JSON.
"""

import json
import openpyxl
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SPREADSHEET = ROOT / "data" / "Dominion - samantekt úr Guðmundi.xlsx"
JSON_FILE = ROOT / "data" / "dominion_data.json"


def build_game_col_map(ws):
    """Map game_num -> column index from the spreadsheet."""
    game_col_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val and isinstance(val, (int, float)) and 1 <= val <= 200:
            game_col_map[int(val)] = col
    return game_col_map


def read_spreadsheet_results(ws, col):
    """Read name↔place pairs from a game column. Names at col, places at col+2."""
    pairs = []
    for row in range(3, 15):  # max 12 players
        name = ws.cell(row=row, column=col).value
        place = ws.cell(row=row, column=col + 2).value
        if not name or not isinstance(name, str):
            continue
        if place is not None and isinstance(place, (int, float)):
            pairs.append({"name": name, "place": int(place)})
        else:
            # Some early games have no places recorded
            pairs.append({"name": name, "place": None})
    return pairs


def main():
    wb = openpyxl.load_workbook(str(SPREADSHEET), data_only=True)
    ws = wb["Öll spil"]
    game_col_map = build_game_col_map(ws)

    with open(JSON_FILE) as f:
        data = json.load(f)

    legacy_games = data.get("legacy_games", [])
    legacy_by_num = {g["game_num"]: g for g in legacy_games if g.get("game_num")}

    fixed = 0
    no_places = 0
    not_found = 0

    for game_num in range(1, 50):
        if game_num not in legacy_by_num:
            print(f"  Game #{game_num}: not in legacy JSON, skipping")
            not_found += 1
            continue

        if game_num not in game_col_map:
            print(f"  Game #{game_num}: not in spreadsheet, skipping")
            not_found += 1
            continue

        col = game_col_map[game_num]
        ss_results = read_spreadsheet_results(ws, col)

        if not ss_results:
            print(f"  Game #{game_num}: no results in spreadsheet")
            continue

        # Check if any places are recorded
        has_places = any(r["place"] is not None for r in ss_results)
        if not has_places:
            # Keep existing JSON results for games without places in spreadsheet
            print(f"  Game #{game_num}: no places in spreadsheet, keeping existing")
            no_places += 1
            continue

        # Build corrected results, preserving score from existing JSON if present
        old_results = {r["name"]: r for r in legacy_by_num[game_num].get("results", [])}
        new_results = []
        for r in ss_results:
            if r["place"] is None:
                continue
            entry = {"name": r["name"], "place": r["place"]}
            # Preserve score if it existed in the old data
            if r["name"] in old_results and old_results[r["name"]].get("score") is not None:
                entry["score"] = old_results[r["name"]]["score"]
            else:
                entry["score"] = None
            new_results.append(entry)

        # Sort by place
        new_results.sort(key=lambda x: x["place"])

        old_set = {(r["name"], r["place"]) for r in legacy_by_num[game_num].get("results", [])}
        new_set = {(r["name"], r["place"]) for r in new_results}

        if old_set != new_set:
            old_sorted = sorted(legacy_by_num[game_num].get("results", []), key=lambda x: x.get("place", 99))
            print(f"  Game #{game_num}: FIXED")
            print(f"    Old: {[(r['name'], r['place']) for r in old_sorted]}")
            print(f"    New: {[(r['name'], r['place']) for r in new_results]}")
            fixed += 1

        # Update the results (also update players list to match)
        legacy_by_num[game_num]["results"] = new_results
        legacy_by_num[game_num]["players"] = [r["name"] for r in new_results]

    print(f"\nSummary: {fixed} games fixed, {no_places} without places, {not_found} not found")

    # Write back
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Updated {JSON_FILE}")


if __name__ == "__main__":
    main()
